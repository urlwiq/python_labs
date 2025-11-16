import csv
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None

def csv_to_xlsx(csv_path: str, xlsx_path: str) -> None:
    csv_file = Path(csv_path)
    
    if not csv_file.exists():
        raise FileNotFoundError(f"CSV файл не найден: {csv_path}")
    
    if csv_file.suffix.lower() != '.csv':
        raise ValueError(f"Неверный тип файла: ожидается .csv, получен {csv_file.suffix}")
    
    if openpyxl is None:
        raise ImportError("Не установлена библиотека openpyxl")
    
    try:
        with open(csv_path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            data = list(reader)
    except Exception as e:
        raise ValueError(f"Ошибка чтения CSV файла: {e}")
    
    if not data:
        raise ValueError("CSV файл пуст")
    

    for row_idx, row in enumerate(data, 1):
        for col_idx, value in enumerate(row, 1):
            sheet.cell(row=row_idx, column=col_idx, value=value)
    header_font = Font(bold=True)
    for cell in sheet[1]:
        cell.font = header_font
    for column_cells in sheet.columns:
        length = max(8, *(
            len(str(cell.value)) if cell.value else 0
            for cell in column_cells
        ))
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2
    
    Path(xlsx_path).parent.mkdir(parents=True, exist_ok=True)
    workbook.save(xlsx_path)

def main() -> None:
    try:
        Path("data/out").mkdir(parents=True, exist_ok=True)
        csv_to_xlsx("data/samples/people.csv", "data/out/people.xlsx")
        print("CSV -> XLSX: Успешно")
        
    except (FileNotFoundError, ValueError, ImportError) as e:
        print(f"Ошибка: {e}")
    except Exception as e:
        print(f"Неожиданная ошибка: {e}")

if __name__ == "__main__":
    main()